import json
import requests
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dotenv import load_dotenv
import re

# Load environment variables (DEEPL API key)
load_dotenv()
DEEPL_API_KEY = os.getenv("DEEPL_FREE_SECRET")

# --- HARDCODED PATHS ---
EXCEL_OLD_TRANSLATIONS = "/Users/wardvercruyssen/Downloads/translations.xlsx"
TRANSLATION_SETS = {
    "1": {
        "label": "MobilityPlus Portal",
        "paths": {
            "EN": "/Users/wardvercruyssen/repos/react_dashboard/src/i18n/en.json",
            "FR": "/Users/wardvercruyssen/repos/react_dashboard/src/i18n/fr.json",
            "NL": "/Users/wardvercruyssen/repos/react_dashboard/src/i18n/nl.json",
        },
    },
    "2": {
        "label": "Synkee Portal",
        "paths": {
            "EN": "/Users/wardvercruyssen/repos/synkee_portal/src/i18n/en.json",
            "FR": "/Users/wardvercruyssen/repos/synkee_portal/src/i18n/fr.json",
            "NL": "/Users/wardvercruyssen/repos/synkee_portal/src/i18n/nl.json",
        },
    },
}

# --- HELPER FUNCTIONS ---

def timestamp():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_excel_translations(path):
    wb = load_workbook(path.strip())
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    lang_map = {col: header.index(col) for col in ["Key", "EN", "FR", "NL"]}
    translations = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[lang_map["Key"]]
        translations[key] = {
            "EN": row[lang_map["EN"]] or "",
            "FR": row[lang_map["FR"]] or "",
            "NL": row[lang_map["NL"]] or "",
        }
    return translations

def write_excel(translations_dict, path, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["Key", "EN", "FR", "NL"])
    for key in sorted(translations_dict.keys()):
        row = [
            key,
            translations_dict[key].get("EN", ""),
            translations_dict[key].get("FR", ""),
            translations_dict[key].get("NL", "")
        ]
        ws.append(row)
    wb.save(path)
    print(f"✅ Wrote: {path} ({len(translations_dict)} keys)")

def protect_placeholders(text):
    pattern = re.compile(r"{{.*?}}|<.*?>.*?</.*?>")
    matches = pattern.findall(text)
    replacements = {}
    for i, match in enumerate(matches):
        token = f"@@{i}@@"
        replacements[token] = match
        text = text.replace(match, token, 1)
    return text, replacements

def restore_placeholders(translated_text, replacements):
    for token, original in replacements.items():
        translated_text = translated_text.replace(token, original)
    return translated_text

def translate_text(text, target_lang, failed_keys=None, key_name=None):
    print(f"🌍 Translating key '{key_name}' to {target_lang}: \"{text}\"")
    safe_text, placeholders = protect_placeholders(text)
    try:
        url = "https://api-free.deepl.com/v2/translate"
        params = {
            "auth_key": DEEPL_API_KEY,
            "text": safe_text,
            "target_lang": target_lang
        }
        response = requests.post(url, data=params)
        response.raise_for_status()
        translated = response.json()["translations"][0]["text"]
        final = restore_placeholders(translated, placeholders)
        print(f"✅ [{target_lang}] {key_name}: \"{final}\"\n")
        return final
    except Exception as e:
        print(f"❌ DeepL translation failed for key '{key_name}' ({target_lang}): {e}")
        if failed_keys is not None and key_name:
            failed_keys.append((key_name, target_lang))
        return ""

# --- MAIN SCRIPT ---
def main():
    print("🚀 Starting translation validation and export...")

    # Ask user which translation set to use
    print("📁 Which translation set do you want to use?")
    for key, value in TRANSLATION_SETS.items():
        print(f" {key}. {value['label']}")

    selection = input("Enter number: ").strip()
    if selection not in TRANSLATION_SETS:
        print("❌ Invalid selection.")
        return

    JSON_PATHS = TRANSLATION_SETS[selection]["paths"]

    # Step 1: Load JSONs and collect all keys
    flattened_jsons = {lang: load_json(path) for lang, path in JSON_PATHS.items()}
    all_keys = set().union(*[set(flattened_jsons[lang].keys()) for lang in JSON_PATHS])

    # Step 2: Enforce EN key existence if key is in FR or NL
    keys_missing_en = [
        key for key in sorted(all_keys)
        if key not in flattened_jsons["EN"]
        and (key in flattened_jsons["FR"] or key in flattened_jsons["NL"])
    ]

    if keys_missing_en:
        print("❌ ERROR: The following key(s) exist in FR/NL but are missing in EN.json:\n")
        for key in keys_missing_en:
            print(f" - {key}")
        print("\n💡 Please add these key(s) to EN.json manually before continuing.")
        return

    # Step 3: Detect missing FR/NL keys and prompt to translate
    validation_errors = []
    for key in sorted(all_keys):
        missing_langs = [lang for lang in ["EN", "FR", "NL"] if key not in flattened_jsons[lang]]
        if missing_langs:
            validation_errors.append((key, missing_langs))

    failed_translations = []
    if validation_errors:
        print("❗ Missing keys detected across JSONs:")
        for key, langs in validation_errors:
            print(f" - Key '{key}' is missing in: {', '.join(langs)}")

        translate_prompt = input("\n💬 Translate missing FR/NL keys with DeepL and update JSONs? (y/n): ").strip().lower()
        if translate_prompt == "y":
            for key, langs in validation_errors:
                en_text = flattened_jsons["EN"].get(key, "")
                for lang in langs:
                    if lang in ["FR", "NL"] and en_text:
                        translated = translate_text(en_text, lang, failed_translations, key)
                        if translated:
                            flattened_jsons[lang][key] = translated
            for lang in ["FR", "NL"]:
                save_json(JSON_PATHS[lang], flattened_jsons[lang])
            print("✅ Missing translations added to JSON files.")
        else:
            print("⚠️ DeepL translation skipped.")

    # Step 4: Abort if any translations still failed or keys are missing
    if failed_translations or any(
        key not in flattened_jsons["EN"] or
        key not in flattened_jsons["FR"] or
        key not in flattened_jsons["NL"]
        for key in all_keys
    ):
        print("\n❌ Some keys are still missing in one or more JSON files.")
        print("💡 Please fix them manually before exporting.")
        return

    # Step 5: Load Excel if it exists
    if os.path.exists(EXCEL_OLD_TRANSLATIONS):
        excel_data = load_excel_translations(EXCEL_OLD_TRANSLATIONS)
        old_keys = set(excel_data.keys())
    else:
        print(f"⚠️ Excel file not found at {EXCEL_OLD_TRANSLATIONS}")
        confirm = input("Proceed assuming ALL keys are new? (y/n): ").strip().lower()
        if confirm != "y":
            print("❌ Aborted.")
            return
        old_keys = set()

    # Step 6: Prepare data for export
    full_translation_set = {
        key: {
            "EN": flattened_jsons["EN"][key],
            "FR": flattened_jsons["FR"][key],
            "NL": flattened_jsons["NL"][key]
        }
        for key in all_keys
    }
    new_keys_only = {
        key: full_translation_set[key] for key in all_keys if key not in old_keys
    }

    # Step 7: Export Excel files
    ts = timestamp()
    write_excel(new_keys_only, f"/Users/wardvercruyssen/Downloads/new-keys-ONLY-{ts}.xlsx", "New Keys Only")
    write_excel(full_translation_set, f"/Users/wardvercruyssen/Downloads/all-keys-MERGED-{ts}.xlsx", "All Keys Merged")

    print("🎉 Done!")

if __name__ == "__main__":
    main()
