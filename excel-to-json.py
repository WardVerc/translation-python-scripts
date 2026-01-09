import json
import os
from openpyxl import load_workbook
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables
load_dotenv()

# --- HARDCODED PATHS ---
EXCEL_REVIEWED_PATH = "./translations-reviewed.xlsx"
TRANSLATION_SETS = {
    "1": {
        "label": "Translation",
        "paths": {
            "EN": "./en.json",
            "FR": "./fr.json",
            "NL": "./nl.json",
        },
    }
}

# --- HELPER FUNCTIONS ---
def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    lang_map = {col: header.index(col) for col in ["Key", "EN", "FR", "NL"]}

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[lang_map["Key"]]
        data[key] = {
            "EN": row[lang_map["EN"]] or "",
            "FR": row[lang_map["FR"]] or "",
            "NL": row[lang_map["NL"]] or ""
        }
    return data

def timestamp():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def write_excel(translations_dict, path, title):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["Key", "EN", "FR", "NL"])
    for key in sorted(translations_dict.keys()):
        row = [
            key,
            translations_dict[key].get("EN", ""),
            translations_dict[key].get("FR", ""),
            translations_dict[key].get("NL", "")
        ]
        ws.append(row)
    wb.save(path)
    print(f"✅ Wrote: {path} ({len(translations_dict)} keys)")

# --- MAIN SCRIPT ---
def main():
    print("🚀 Starting import of reviewed translations...")

    if not EXCEL_REVIEWED_PATH.endswith("translations-reviewed.xlsx"):
        print("❌ The reviewed Excel file must be named 'translations-reviewed.xlsx'. Rename it and try again.")
        return

    if not os.path.exists(EXCEL_REVIEWED_PATH):
        print(f"❌ File not found: {EXCEL_REVIEWED_PATH}")
        print("💡 Please make sure the file exists and is named correctly.")
        return

    # Ask which translation set to use
    print("📁 Which translation set do you want to use?")
    for key, value in TRANSLATION_SETS.items():
        print(f" {key}. {value['label']}")
    selection = input("Enter number: ").strip()
    if selection not in TRANSLATION_SETS:
        print("❌ Invalid selection.")
        return

    JSON_PATHS = TRANSLATION_SETS[selection]["paths"]

    reviewed_data = load_excel(EXCEL_REVIEWED_PATH)
    jsons = {lang: load_json(path) for lang, path in JSON_PATHS.items()}

    changes = {"EN": {}, "FR": {}, "NL": {}}

    for key, new_vals in reviewed_data.items():
        for lang in ["EN", "FR", "NL"]:
            old_val = jsons[lang].get(key, "")
            new_val = new_vals[lang]
            if old_val != new_val and new_val:
                changes[lang][key] = new_val

    for lang in ["EN", "FR", "NL"]:
        if changes[lang]:
            print(f"\n🔄 {lang} - {len(changes[lang])} updates:")
            for key, val in changes[lang].items():
                print(f" - {key}: '{val}'")

    confirm = input("\n💾 Apply these changes to the JSON files? (y/n): ").strip().lower()
    if confirm != "y":
        print("❌ Operation cancelled.")
        return

    for lang in ["EN", "FR", "NL"]:
        jsons[lang].update(changes[lang])
        save_json(JSON_PATHS[lang], jsons[lang])

    all_keys = set().union(*[jsons[lang].keys() for lang in ["EN", "FR", "NL"]])
    full_export = {
        key: {
            "EN": jsons["EN"].get(key, ""),
            "FR": jsons["FR"].get(key, ""),
            "NL": jsons["NL"].get(key, "")
        }
        for key in all_keys
    }
    write_excel(full_export, f"./all-keys-MERGED-{timestamp()}.xlsx", "All Keys Merged")

    print("\n🎉 Import completed!")

if __name__ == "__main__":
    main()
